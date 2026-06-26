import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_to_excel(registrations, stats):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Registrations ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "التسجيلات"
    ws.sheet_view.rightToLeft = True

    H_FONT   = Font(bold=True, color="FFFFFF", size=11)
    H_FILL   = PatternFill("solid", fgColor="1E3A5F")
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
    CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP     = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="BBBBBB")
    BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "م", "لقب الطالب 1", "اسم الطالب 1", "بريد الطالب 1",
        "لقب الطالب 2", "اسم الطالب 2", "بريد الطالب 2",
        "نوع التسجيل", "التخصص", "عنوان المذكرة", "المشرف",
        "ملاحظات", "طلب التغيير", "تاريخ التسجيل",
    ]
    widths = [5, 18, 18, 32, 18, 18, 32, 14, 28, 55, 28, 35, 45, 20]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = H_FONT, H_FILL, CENTER, BORDER
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for i, reg in enumerate(registrations, 1):
        title = reg.get("title_name") or reg.get("custom_title") or ""
        sup   = reg.get("supervisor_name") or reg.get("custom_supervisor") or ""
        fill  = ALT_FILL if i % 2 == 0 else None

        row_data = [
            i,
            reg["student1_last"], reg["student1_first"], reg["student1_email"],
            reg.get("student2_last") or "", reg.get("student2_first") or "",
            reg.get("student2_email") or "",
            "فردي" if reg["is_solo"] else "ثنائي",
            reg.get("spec_name") or "", title, sup,
            reg.get("notes") or "", reg.get("change_request") or "",
            reg.get("created_at", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = CENTER if col in (1, 8) else WRAP
            cell.border = BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[i + 1].height = 28

    # ── Sheet 2: Stats ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("إحصائيات")
    ws2.sheet_view.rightToLeft = True
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 18

    stat_rows = [
        ("إجمالي التسجيلات",          stats["total"]),
        ("تسجيلات فردية",              stats["solo"]),
        ("مقاعد فردية متبقية",         stats["solo_remaining"]),
        ("تسجيلات ثنائية",             stats["duo"]),
        ("طلبات التغيير",              stats["change_requests"]),
    ]
    BLD = Font(bold=True, size=11)
    for row, (label, val) in enumerate(stat_rows, 1):
        ws2.cell(row=row, column=1, value=label).font = BLD
        c = ws2.cell(row=row, column=2, value=val)
        c.alignment = Alignment(horizontal="center")

    # Supervisors table
    ws2.cell(row=len(stat_rows)+2, column=1, value="المشرف").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=len(stat_rows)+2, column=1).fill = PatternFill("solid", fgColor="1E3A5F")
    ws2.cell(row=len(stat_rows)+2, column=2, value="المذكرات الحالية / الحد الأقصى").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=len(stat_rows)+2, column=2).fill = PatternFill("solid", fgColor="1E3A5F")

    for j, sup in enumerate(stats["supervisors"], len(stat_rows)+3):
        ws2.cell(row=j, column=1, value=sup["name"])
        ws2.cell(row=j, column=2, value=f"{sup['current_count']} / {sup['max_students']}")
        ws2.cell(row=j, column=2).alignment = Alignment(horizontal="center")

    # ── Output ─────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
